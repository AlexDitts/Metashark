import json
from typing import List
import time
from education.models import Direction, StudyGroup
import openpyxl
from django.db.models import Q, Count


def get_direction_data() -> List[dict]:
    """
    Функция возвращает список из словарей где указаны название направлений, данные кураторов и списки дисциплин
    по каждому направлению
    :return: List[dict]
    """
    data_directions = []
    for direct in Direction.objects.all():
        item_direction = {'direction': direct.title,
                          'name': direct.curator.user.first_name,
                          'last_name': direct.curator.user.last_name,
                          'email': direct.curator.user.email,
                          'phone': direct.curator.phone,
                          'list_discipline': direct.discipline.all().values('title')
                          }
        data_directions.append(item_direction)

    return data_directions


def get_groups_data() -> List[dict]:
    """
    Функция возвращает список из словарей где указаны номера групп, количество мужчин и женщин в группах,
    количество свободных мест в группе, а так же отсортированный список с именами студентов в каждой группе
    :return: List[dict]
    """
    groups_queryset = StudyGroup.objects.annotate(
        vacancies=20 - Count('student'),
        males=Count('student', filter=Q(student__gender=0)),
        females=Count('student', filter=Q(student__gender=1))
    )
    data_group = []
    for group in groups_queryset:
        item_group = {'group_number': group.number,
                      'number_of_men': group.males,
                      'number_of_women': group.females,
                      'vacancies': group.vacancies,
                      'students_list': list(group.student.all().order_by('name').values('name'))
                      }

        data_group.append(item_group)
    return data_group


def report_to_exel() -> None:
    """
    Функция создаёт отчёт по направлениям и группам и создаёт файл эксель, куда заносит все данные из отчёта.
    :return: None
    """
    directions = get_direction_data()
    groups = get_groups_data()
    indent = 2

    book = openpyxl.Workbook()
    sheet = book.active

    first_title = list(directions[0].keys())
    for column, item in enumerate(first_title, 2):
        sheet.cell(indent, column).value = item

    for row, direct in enumerate(directions, indent):
        sheet.cell(row + indent, 2).value = direct['direction']
        sheet.cell(row + indent, 3).value = direct['name']
        sheet.cell(row + indent, 4).value = direct['last_name']
        sheet.cell(row + indent, 5).value = direct['email']
        sheet.cell(row + indent, 6).value = direct['phone']
        list_discipline = direct['list_discipline']

        for inter_row, discipline in enumerate(list_discipline, row):
            sheet.cell(inter_row + indent, 7).value = discipline['title']
        indent += len(list_discipline)

    indent += len(directions) + 4

    title_report = list(groups[0].keys())

    for column, item in enumerate(title_report, 2):
        sheet.cell(indent, column).value = item

    indent += 2
    for row, group in enumerate(groups):
        sheet.cell(row + indent, 2).value = group['group_number']
        sheet.cell(row + indent, 3).value = group['number_of_men']
        sheet.cell(row + indent, 4).value = group['number_of_women']
        sheet.cell(row + indent, 5).value = group['vacancies']
        students_list = group['students_list']
        for inter_row, student in enumerate(students_list, row):
            sheet.cell(inter_row + indent, 6).value = student['name']
        indent += len(students_list) + 1

    current_time = time.strftime('%H-%M-%S')
    book.save(f'{current_time}_test_file.xlsx')
    book.close()
